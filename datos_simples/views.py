from .serializers import CursoSerializer, CarreraSerializer, DistribucionCargaAcademicaSerializer, DocenteSerializer
from .models import Curso, Carrera, DistribucionCargaAcademica, Docente

from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.viewsets import ModelViewSet
from rest_framework.decorators import api_view
from django.http import JsonResponse

import pandas as pd
from openpyxl import load_workbook
from sqlalchemy import create_engine

# Motor de postgres
url = 'postgresql+psycopg2://roswell:12345@localhost:5432/ingenieria_software'
engine = create_engine(url=url, echo=False)

# Funciones utilitarias
def retrieveAll(modelo):
  todo = modelo.objects.all()
  return todo

split_data = lambda doble_curso : doble_curso.split(' / ')

def agregarFilas(atributo):
  return pd.Series(map(split_data, atributo))

def atributoADataframe(atributo, nombre):
  dataframe = pd.DataFrame(atributo, columns=[nombre])
  return dataframe

def guardarUsuario(usuario):
  user = Docente(nombre=usuario, staff=True, admin=False)
  user.set_password(usuario)
  user.save()

def agregarUsuarios(usuarios):
  pd.Series(map(guardarUsuario, usuarios))

# Create your views here.
class CursoViewSet(ModelViewSet):
  queryset = Curso.objects.all()
  serializer_class = CursoSerializer

class CarreraViewSet(ModelViewSet):
  queryset = Carrera.objects.all()
  serializer_class = CarreraSerializer

@api_view(['GET'])
def view_adocente(request):
  archivos = retrieveAll(DistribucionCargaAcademica)
  archivo = archivos[0]
  archivo = archivo.archivo
  # leer el workbook
  data_excel = load_workbook(archivo, data_only=True)
  hoja_distribucion = data_excel['DISTR. SEM 2022-2-INF']
  data = hoja_distribucion.tables['CARGA_ACAD']
  data = hoja_distribucion[data.ref]
  lista_filas = []
  for fila in data:
    atributos = []
    for atributo in fila:
      atributos.append(atributo.value)
    lista_filas.append(atributos)
  dataframe = pd.DataFrame(data=lista_filas[1:], index=None, columns=lista_filas[0])
  agregarDocentes(dataframe['DOCENTES'].unique())
  return JsonResponse({'msg': 'correcto'})


def agregarDocentes(data):
  docentes = atributoADataframe(data, 'docente')
  index_inactivos = docentes[ (docentes['docente'] == 'GRUPO DESACTIVADO') | (docentes['docente'] == 'CURSO O GRUPO POR DESACTIVAR') | (docentes['docente'] == 'CURSO DESACTIVADO')].index
  docentes.drop(index_inactivos, inplace=True)
  agregarUsuarios(docentes['docente'])

class DistribucionCargaAcademicaViewSet(ModelViewSet):
  queryset = DistribucionCargaAcademica.objects.all()
  serializer_class = DistribucionCargaAcademicaSerializer
  parser_classes = (MultiPartParser, FormParser)

  def perform_create(self, serializer):
    archivos = DistribucionCargaAcademica.objects.all()
    cantidad = len(archivos)
    if cantidad == 0:
      serializer.save()
      archivos = retrieveAll(DistribucionCargaAcademica)
      archivo = archivos[0]
      archivo = archivo.archivo
      # leer el workbook
      data_excel = load_workbook(archivo, data_only=True)
      hoja_distribucion = data_excel['DISTR. SEM 2022-2-INF']
      data = hoja_distribucion.tables['CARGA_ACAD']
      data = hoja_distribucion[data.ref]
      lista_filas = []
      for fila in data:
        atributos = []
        for atributo in fila:
          atributos.append(atributo.value)
        lista_filas.append(atributos)
      dataframe = pd.DataFrame(data=lista_filas[1:], index=None, columns=lista_filas[0])
      dataframe['CARRERA'] = agregarFilas(dataframe['CARRERA'])
      dataframe['CODIGO'] = agregarFilas(dataframe['CODIGO'])
      dataframe = dataframe.explode(['CODIGO', 'CARRERA'])
      dataframe.drop(['N°'], axis=1, inplace=True)
      # recuperar cursos
      cursos = atributoADataframe(dataframe['CURSO'].unique(), 'nombre')
      # recuperar carreras
      carreras = atributoADataframe(dataframe['CARRERA'].unique(), 'nombre')
      # agregar cursos y carreras
      with engine.connect().execution_options(autocommit=True) as conn:
        cursos.to_sql(Curso._meta.db_table, con=conn, index=False, if_exists='append', method='multi')
        carreras.to_sql(Carrera._meta.db_table, con=conn, index=False, if_exists='append', method='multi')
      # recuperar docentes
      docentes = atributoADataframe(dataframe['DOCENTES'].unique(), 'docente')
      # borrar nombres que no son docentes
      index_inactivo = docentes[ docentes['docente'] == 'GRUPO DESACTIVADO' ].index
      docentes.drop(index_inactivo, inplace=True)
      # agregar docentes
      agregarUsuarios(docentes['docente'])
      # agregar cursos activos
      # -- recuperar 3 atributos
      cursos_activos = dataframe[['CODIGO', 'CURSO', 'CARRERA']]
      cursos_activos.drop_duplicates(inplace=True)
      # -- recuperar los cursos de la base de datos
      cursos_db = retrieveAll(Curso)
      cursos_db = atributoADataframe(cursos_db, 'objecto_curso')
      # -- recuperar las carreras de la base de datos
      carreras_db = retrieveAll(Carrera)
      carreras_db = atributoADataframe(carreras_db, 'objeto_carrera')
      # agregar horarios



class DocenteViewSet(ModelViewSet):
  queryset = Docente.objects.all()
  serializer_class = DocenteSerializer